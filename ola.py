import streamlit as st
import psycopg2
from psycopg2 import pool
import pandas as pd
from datetime import datetime, timedelta, time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz
from contextlib import contextmanager

# ==================== CONFIGURACIÓN ====================
st.set_page_config(
    page_title="JP Transportamos",
    layout="wide",
    page_icon="🚛",
    initial_sidebar_state="collapsed"
)

# ==================== CREDENCIALES ====================
SUPABASE_DB_URL = "postgresql://postgres.verwlkgitpllyneqxlao:Conejito800$@aws-0-us-west-2.pooler.supabase.com:6543/postgres"

# ==================== CATÁLOGO PLACAS / CONDUCTORES ====================
PLACA_CONDUCTOR = {
    "WCP116": "JOSE MIGUEL",
    "SKL440": "JESUS GABRIEL",
    "SVB345": "ALBERTO DE LA CRUZ RONDERO",
    "VEB931": "ALVARO ANTONIO",
    "XIC344": "NILSON BERNAL",
    "SMD925": "LUIS ECHEVERRY",
    "SRE493": "GUILERMO ENRIQUE",
    "VAG227": "GIOVANNY JUNIOR",
    "SOD747": "YOHANIS  JESUS",
}

TODOS_CONDUCTORES = sorted([
    "JOSE MIGUEL", "JESUS GABRIEL", "ALBERTO DE LA CRUZ RONDERO", "ALVARO ANTONIO",
    "NILSON BERNAL", "LUIS ECHEVERRY", "GUILERMO ENRIQUE", "GIOVANNY JUNIOR", "YOHANIS  JESUS"
])

ESTADOS_VIAJE = ["✅ Completado", "❌ Anulado", "⚠️ Incumplido", "🔄 En Curso"]

# ==================== RUTAS FRECUENTES ====================
RUTAS_FRECUENTES = [
    ("ORIGEN A", "DESTINO A"),
    ("ORIGEN B", "DESTINO B"),
]

ORIGENES_FRECUENTES = sorted(set(r[0] for r in RUTAS_FRECUENTES))
LABEL_MANUAL = "✏️ Escribir manualmente..."

# ==================== CLIENTES FRECUENTES ====================
CLIENTES_FRECUENTES = [
    "CLIENTE 1",
    "CLIENTE 2",
]
LABEL_MANUAL_CLI = "✏️ Escribir manualmente..."

# ==================== COORDENADAS ====================
COORDENADAS = {
    "ORIGEN A": (4.6097, -74.0817),
    "DESTINO A": (4.7110, -74.0721),
}

# ==================== BONO AUTOMÁTICO POR HORA DE CITA ====================
def calcular_bono_transporte(hora_cita) -> int:
    if hora_cita is None:
        return 0
    if isinstance(hora_cita, str):
        try:
            parts = hora_cita[:5].split(":")
            hora_cita = time(int(parts[0]), int(parts[1]))
        except:
            return 0
    if not isinstance(hora_cita, time):
        return 0
    h, m = hora_cita.hour, hora_cita.minute
    total_min = h * 60 + m
    if total_min == 360:
        return 10000
    if 1140 <= total_min <= 1260:
        return 10000
    if 1290 <= total_min <= 1380:
        return 20000
    return 0

def fmt_moneda(val) -> str:
    if val is None:
        return "—"
    try:
        return f"${int(val):,}".replace(",", ".")
    except:
        return "—"

def parse_moneda(texto: str) -> int:
    if not texto:
        return 0
    try:
        limpio = texto.strip().replace("$", "").replace(" ", "")
        if "." in limpio and "," not in limpio:
            limpio = limpio.replace(".", "")
        elif "," in limpio:
            limpio = limpio.replace(".", "").replace(",", ".")
        return int(float(limpio))
    except:
        return 0

def input_moneda(label: str, key: str, value: int = 0, help_text: str = None) -> int:
    texto = st.text_input(
        label,
        value=str(value) if value else "",
        placeholder="Ej: 1.200.000",
        key=key,
        help=help_text or "Escribe el valor con o sin puntos: 1200000 o 1.200.000"
    )
    val_int = parse_moneda(texto)
    if texto and val_int > 0:
        st.caption(f"✅ {fmt_moneda(val_int)}")
    elif texto and val_int == 0 and texto.strip() not in ("", "0"):
        st.caption("⚠️ Valor no reconocido")
    return val_int

# ==================== CSS ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Exo+2:wght@300;400;500;600&display=swap');

    html, body, [class*="css"] {
        font-family: 'Exo 2', sans-serif;
    }

    .main-header {
        background: linear-gradient(135deg, #0a0a0a 0%, #1a1a2e 40%, #16213e 70%, #0f3460 100%);
        padding: 1.8rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 1.5rem;
        border: 1px solid #e94560;
        box-shadow: 0 0 30px rgba(233,69,96,0.15), inset 0 1px 0 rgba(255,255,255,0.05);
        position: relative;
        overflow: hidden;
    }
    .main-header::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: radial-gradient(circle, rgba(233,69,96,0.08) 0%, transparent 70%);
        border-radius: 50%;
    }
    .main-header h1 {
        font-family: 'Rajdhani', sans-serif;
        font-size: 2.2rem;
        font-weight: 700;
        color: #ffffff;
        margin: 0;
        letter-spacing: 3px;
        text-transform: uppercase;
    }
    .main-header .subtitle {
        color: #e94560;
        margin: 0.2rem 0 0 0;
        font-size: 0.85rem;
        letter-spacing: 2px;
        text-transform: uppercase;
        font-weight: 500;
    }
    .main-header .empresa-badge {
        display: inline-block;
        background: #e94560;
        color: white;
        padding: 2px 10px;
        border-radius: 4px;
        font-size: 0.7rem;
        font-weight: 700;
        letter-spacing: 2px;
        margin-bottom: 0.5rem;
        text-transform: uppercase;
    }

    .kpi-card {
        background: linear-gradient(135deg, #1a1a2e, #16213e);
        border: 1px solid rgba(233,69,96,0.3);
        border-radius: 12px;
        padding: 1.1rem 1.3rem;
        margin-bottom: 0.5rem;
        box-shadow: 0 4px 15px rgba(0,0,0,0.3);
    }
    .kpi-card .kpi-val {
        font-family: 'Rajdhani', sans-serif;
        font-size: 1.9rem;
        font-weight: 700;
        color: #e94560;
    }
    .kpi-card .kpi-lbl {
        font-size: 0.72rem;
        color: #8892b0;
        text-transform: uppercase;
        letter-spacing: 1.5px;
    }
    .kpi-card.green .kpi-val { color: #2ecc71; }
    .kpi-card.blue  .kpi-val { color: #3498db; }
    .kpi-card.gold  .kpi-val { color: #f39c12; }

    .bono-badge {
        display: inline-block;
        background: linear-gradient(90deg, #f39c12, #e67e22);
        color: white;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: 700;
        font-size: 0.95rem;
        box-shadow: 0 2px 8px rgba(243,156,18,0.4);
    }
    .bono-zero {
        display: inline-block;
        background: #2c3e50;
        color: #8892b0;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 0.9rem;
    }

    .liquidacion-box {
        background: linear-gradient(135deg, #0d1b2a, #1b2838);
        border: 1px solid rgba(46,204,113,0.3);
        border-radius: 12px;
        padding: 1rem 1.5rem;
        margin: 0.8rem 0;
    }
    .liq-row {
        display: flex;
        justify-content: space-between;
        padding: 4px 0;
        font-size: 0.9rem;
        color: #ccd6f6;
        border-bottom: 1px solid rgba(255,255,255,0.05);
    }
    .liq-row:last-child { border-bottom: none; }
    .liq-total {
        display: flex;
        justify-content: space-between;
        padding: 8px 0 0 0;
        font-weight: 700;
        font-size: 1.1rem;
        color: #2ecc71;
        border-top: 1px solid rgba(46,204,113,0.3);
        margin-top: 4px;
    }
    .liq-neg { color: #e74c3c !important; }

    div[data-testid="stTabs"] button {
        font-family: 'Rajdhani', sans-serif;
        font-weight: 600;
        font-size: 1rem;
        letter-spacing: 1px;
        text-transform: uppercase;
    }

    .trazabilidad-header {
        background: linear-gradient(90deg, #1a1a2e, #16213e);
        border-left: 4px solid #e94560;
        padding: 0.6rem 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0 0.5rem 0;
    }
    .trazabilidad-header h4 {
        margin: 0;
        color: #ccd6f6;
        font-family: 'Rajdhani', sans-serif;
        letter-spacing: 1px;
    }

    .combustible-card {
        background: linear-gradient(135deg, #0d1b2a, #1a2a1a);
        border: 1px solid rgba(46,204,113,0.4);
        border-radius: 12px;
        padding: 1rem 1.5rem;
        margin: 0.5rem 0;
    }
    .saldo-alto  { color: #2ecc71; font-weight: 700; font-size: 1.4rem; font-family: 'Rajdhani', sans-serif; }
    .saldo-medio { color: #f39c12; font-weight: 700; font-size: 1.4rem; font-family: 'Rajdhani', sans-serif; }
    .saldo-bajo  { color: #e74c3c; font-weight: 700; font-size: 1.4rem; font-family: 'Rajdhani', sans-serif; }
</style>
""", unsafe_allow_html=True)


# ==================== POOL DE CONEXIONES ====================
# Un solo pool compartido para todos los usuarios — thread-safe
@st.cache_resource(show_spinner=False)
def get_connection_pool():
    return pool.ThreadedConnectionPool(
        minconn=2,
        maxconn=10,
        dsn=SUPABASE_DB_URL
    )


# ==================== BASE DE DATOS ====================
class DB:
    def __init__(self):
        self.url = SUPABASE_DB_URL
        self.init()

    @contextmanager
    def get_conn(self):
        """
        Context manager que obtiene una conexión del pool,
        hace commit al salir o rollback si hay error,
        y devuelve la conexión al pool siempre.
        """
        conn_pool = get_connection_pool()
        c = conn_pool.getconn()
        try:
            yield c
            c.commit()
        except Exception as e:
            c.rollback()
            raise e
        finally:
            conn_pool.putconn(c)

    def init(self):
        if st.session_state.get("_db_init_done"):
            return
        try:
            with self.get_conn() as c:
                cur = c.cursor()
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS jp_viajes (
                        id SERIAL PRIMARY KEY,
                        fecha_registro TIMESTAMP DEFAULT (now() AT TIME ZONE 'America/Bogota'),
                        fecha DATE NOT NULL,
                        placa TEXT NOT NULL,
                        conductor TEXT,
                        cliente TEXT,
                        origen TEXT,
                        destino TEXT,
                        hora_cita_cargue TIME,
                        hora_salida_cargue TIME,
                        hora_llegada_descargue TIME,
                        hora_salida_descargue TIME,
                        contenedor TEXT,
                        carga TEXT,
                        numero_factura TEXT,
                        manifiesto TEXT,
                        observacion TEXT,
                        estado TEXT DEFAULT 'Completado',
                        dias_salida_cargue INTEGER DEFAULT 0,
                        dias_llegada_descargue INTEGER DEFAULT 0,
                        dias_salida_descargue INTEGER DEFAULT 0,
                        flete NUMERIC(12,0) DEFAULT 0,
                        comision NUMERIC(12,0) DEFAULT 0,
                        bono_transporte NUMERIC(12,0) DEFAULT 0,
                        combustible_pesos NUMERIC(12,0) DEFAULT 0,
                        combustible_galones NUMERIC(8,2) DEFAULT 0
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS jp_tanqueos (
                        id SERIAL PRIMARY KEY,
                        fecha_registro TIMESTAMP DEFAULT (now() AT TIME ZONE 'America/Bogota'),
                        fecha DATE NOT NULL,
                        placa TEXT NOT NULL,
                        conductor TEXT,
                        galones NUMERIC(8,2) NOT NULL,
                        costo_pesos NUMERIC(12,0) DEFAULT 0,
                        observacion TEXT
                    )
                """)
                # Migraciones seguras — IF NOT EXISTS evita errores en concurrencia
                migraciones = [
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS flete NUMERIC(12,0) DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS comision NUMERIC(12,0) DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS bono_transporte NUMERIC(12,0) DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS combustible_pesos NUMERIC(12,0) DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS combustible_galones NUMERIC(8,2) DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS numero_factura TEXT",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS dias_salida_cargue INTEGER DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS dias_llegada_descargue INTEGER DEFAULT 0",
                    "ALTER TABLE jp_viajes ADD COLUMN IF NOT EXISTS dias_salida_descargue INTEGER DEFAULT 0",
                ]
                for m in migraciones:
                    try:
                        cur.execute(m)
                    except Exception:
                        pass  # columna ya existe, ignorar
            st.session_state["_db_init_done"] = True
        except Exception as e:
            st.error(f"Error DB init: {e}")

    def guardar_viaje(self, datos: dict) -> bool:
        try:
            with self.get_conn() as c:
                cur = c.cursor()
                cur.execute("""
                    INSERT INTO jp_viajes
                    (fecha, placa, conductor, cliente, origen, destino,
                     hora_cita_cargue, hora_salida_cargue, hora_llegada_descargue, hora_salida_descargue,
                     contenedor, carga, numero_factura, manifiesto, observacion, estado,
                     dias_salida_cargue, dias_llegada_descargue, dias_salida_descargue,
                     flete, comision, bono_transporte, combustible_pesos, combustible_galones)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    datos["fecha"], datos["placa"], datos["conductor"], datos["cliente"],
                    datos["origen"], datos["destino"],
                    datos["hora_cita_cargue"], datos["hora_salida_cargue"],
                    datos["hora_llegada_descargue"], datos["hora_salida_descargue"],
                    datos["contenedor"], datos["carga"],
                    datos["numero_factura"], datos["manifiesto"],
                    datos["observacion"], datos["estado"],
                    datos.get("dias_salida_cargue", 0),
                    datos.get("dias_llegada_descargue", 0),
                    datos.get("dias_salida_descargue", 0),
                    datos.get("flete", 0),
                    datos.get("comision", 0),
                    datos.get("bono_transporte", 0),
                    datos.get("combustible_pesos", 0),
                    datos.get("combustible_galones", 0),
                ))
            return True
        except Exception as e:
            st.error(f"Error guardando: {e}")
            return False

    def actualizar_viaje(self, viaje_id: int, datos: dict) -> bool:
        try:
            with self.get_conn() as c:
                cur = c.cursor()
                cur.execute("""
                    UPDATE jp_viajes SET
                    fecha=%s, placa=%s, conductor=%s, cliente=%s, origen=%s, destino=%s,
                    hora_cita_cargue=%s, hora_salida_cargue=%s,
                    hora_llegada_descargue=%s, hora_salida_descargue=%s,
                    contenedor=%s, carga=%s, numero_factura=%s,
                    manifiesto=%s, observacion=%s, estado=%s,
                    dias_salida_cargue=%s, dias_llegada_descargue=%s, dias_salida_descargue=%s,
                    flete=%s, comision=%s, bono_transporte=%s, combustible_pesos=%s, combustible_galones=%s
                    WHERE id=%s
                """, (
                    datos["fecha"], datos["placa"], datos["conductor"], datos["cliente"],
                    datos["origen"], datos["destino"],
                    datos["hora_cita_cargue"], datos["hora_salida_cargue"],
                    datos["hora_llegada_descargue"], datos["hora_salida_descargue"],
                    datos["contenedor"], datos["carga"],
                    datos["numero_factura"], datos["manifiesto"],
                    datos["observacion"], datos["estado"],
                    datos.get("dias_salida_cargue", 0),
                    datos.get("dias_llegada_descargue", 0),
                    datos.get("dias_salida_descargue", 0),
                    datos.get("flete", 0),
                    datos.get("comision", 0),
                    datos.get("bono_transporte", 0),
                    datos.get("combustible_pesos", 0),
                    datos.get("combustible_galones", 0),
                    viaje_id
                ))
            return True
        except Exception as e:
            st.error(f"Error actualizando: {e}")
            return False

    def eliminar_viaje(self, viaje_id: int) -> bool:
        try:
            with self.get_conn() as c:
                cur = c.cursor()
                cur.execute("DELETE FROM jp_viajes WHERE id=%s", (viaje_id,))
            return True
        except Exception as e:
            st.error(f"Error eliminando: {e}")
            return False

    def obtener_viajes(self, fecha_ini=None, fecha_fin=None, placa=None,
                       conductor=None, cliente=None, estado=None,
                       manifiesto=None, numero_factura=None) -> pd.DataFrame:
        q = """SELECT id, fecha, placa, conductor, cliente, origen, destino,
                      hora_cita_cargue, hora_salida_cargue,
                      hora_llegada_descargue, hora_salida_descargue,
                      contenedor, carga, numero_factura,
                      manifiesto, observacion, estado,
                      COALESCE(dias_salida_cargue,0) as dias_salida_cargue,
                      COALESCE(dias_llegada_descargue,0) as dias_llegada_descargue,
                      COALESCE(dias_salida_descargue,0) as dias_salida_descargue,
                      COALESCE(flete,0) as flete,
                      COALESCE(comision,0) as comision,
                      COALESCE(bono_transporte,0) as bono_transporte,
                      COALESCE(combustible_pesos,0) as combustible_pesos,
                      COALESCE(combustible_galones,0) as combustible_galones
               FROM jp_viajes WHERE 1=1"""
        params = []
        if fecha_ini: q += " AND fecha >= %s"; params.append(fecha_ini)
        if fecha_fin: q += " AND fecha <= %s"; params.append(fecha_fin)
        if placa and placa != "Todas": q += " AND placa = %s"; params.append(placa)
        if conductor: q += " AND conductor ILIKE %s"; params.append(f"%{conductor}%")
        if cliente: q += " AND cliente ILIKE %s"; params.append(f"%{cliente}%")
        if estado and estado != "Todos": q += " AND estado = %s"; params.append(estado)
        if manifiesto: q += " AND manifiesto ILIKE %s"; params.append(f"%{manifiesto}%")
        if numero_factura: q += " AND numero_factura ILIKE %s"; params.append(f"%{numero_factura}%")
        q += " ORDER BY fecha DESC, id DESC"
        try:
            with self.get_conn() as c:
                return pd.read_sql(q, c, params=params)
        except Exception:
            return pd.DataFrame()

    def stats_dashboard(self, fecha_ini, fecha_fin):
        try:
            with self.get_conn() as c:
                return pd.read_sql("""
                    SELECT fecha, placa, conductor, cliente, estado,
                           hora_cita_cargue, hora_salida_cargue,
                           hora_llegada_descargue, hora_salida_descargue,
                           COALESCE(dias_salida_cargue,0) as dias_salida_cargue,
                           COALESCE(dias_llegada_descargue,0) as dias_llegada_descargue,
                           COALESCE(dias_salida_descargue,0) as dias_salida_descargue,
                           COALESCE(flete,0) as flete,
                           COALESCE(comision,0) as comision,
                           COALESCE(bono_transporte,0) as bono_transporte,
                           COALESCE(combustible_pesos,0) as combustible_pesos,
                           COALESCE(combustible_galones,0) as combustible_galones
                    FROM jp_viajes
                    WHERE fecha >= %s AND fecha <= %s
                    ORDER BY fecha
                """, c, params=[fecha_ini, fecha_fin])
        except Exception:
            return pd.DataFrame()

    # ==================== TANQUEOS ====================
    def guardar_tanqueo(self, datos: dict) -> bool:
        try:
            with self.get_conn() as c:
                cur = c.cursor()
                cur.execute("""
                    INSERT INTO jp_tanqueos (fecha, placa, conductor, galones, costo_pesos, observacion)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (
                    datos["fecha"], datos["placa"], datos["conductor"],
                    datos["galones"], datos["costo_pesos"], datos.get("observacion", "")
                ))
            return True
        except Exception as e:
            st.error(f"Error guardando tanqueo: {e}")
            return False

    def eliminar_tanqueo(self, tanqueo_id: int) -> bool:
        try:
            with self.get_conn() as c:
                cur = c.cursor()
                cur.execute("DELETE FROM jp_tanqueos WHERE id=%s", (tanqueo_id,))
            return True
        except Exception as e:
            st.error(f"Error eliminando tanqueo: {e}")
            return False

    def obtener_tanqueos(self, placa=None, fecha_ini=None, fecha_fin=None) -> pd.DataFrame:
        q = "SELECT * FROM jp_tanqueos WHERE 1=1"
        params = []
        if placa and placa != "Todas": q += " AND placa = %s"; params.append(placa)
        if fecha_ini: q += " AND fecha >= %s"; params.append(fecha_ini)
        if fecha_fin: q += " AND fecha <= %s"; params.append(fecha_fin)
        q += " ORDER BY fecha DESC, id DESC"
        try:
            with self.get_conn() as c:
                return pd.read_sql(q, c, params=params)
        except Exception:
            return pd.DataFrame()

    def saldo_combustible_placa(self, placa: str) -> dict:
        try:
            with self.get_conn() as c:
                df_t = pd.read_sql(
                    "SELECT fecha, galones, costo_pesos, 'TANQUEO' as tipo, observacion FROM jp_tanqueos WHERE placa=%s ORDER BY fecha, id",
                    c, params=[placa]
                )
                df_v = pd.read_sql(
                    "SELECT fecha, COALESCE(combustible_galones,0) as galones, COALESCE(combustible_pesos,0) as costo_pesos, 'CONSUMO' as tipo, CONCAT('Viaje ', CAST(id AS TEXT)) as observacion FROM jp_viajes WHERE placa=%s AND COALESCE(combustible_galones,0)>0 ORDER BY fecha, id",
                    c, params=[placa]
                )
            total_tanqueado = float(df_t["galones"].sum()) if not df_t.empty else 0.0
            total_consumido = float(df_v["galones"].sum()) if not df_v.empty else 0.0
            saldo = total_tanqueado - total_consumido

            if not df_t.empty or not df_v.empty:
                df_hist = pd.concat([df_t, df_v], ignore_index=True)
                df_hist = df_hist.sort_values("fecha").reset_index(drop=True)
                saldo_acum = []
                acum = 0.0
                for _, row in df_hist.iterrows():
                    if row["tipo"] == "TANQUEO":
                        acum += float(row["galones"])
                    else:
                        acum -= float(row["galones"])
                    saldo_acum.append(round(acum, 2))
                df_hist["saldo_acumulado"] = saldo_acum
            else:
                df_hist = pd.DataFrame()

            return {
                "total_tanqueado": round(total_tanqueado, 2),
                "total_consumido": round(total_consumido, 2),
                "saldo": round(saldo, 2),
                "historial": df_hist
            }
        except Exception:
            return {"total_tanqueado": 0, "total_consumido": 0, "saldo": 0, "historial": pd.DataFrame()}


# ==================== INSTANCIA DB POR SESIÓN ====================
# Cada usuario tiene su propia instancia en session_state —
# evita que una sesión interfiera con otra y no hay caché global de conexiones.
def get_db() -> DB:
    if "db_instance" not in st.session_state:
        st.session_state["db_instance"] = DB()
    return st.session_state["db_instance"]


# ==================== CACHÉ DE CONSULTAS ====================
# TTL corto (30s) para que los datos se refresquen rápido entre usuarios.
# El underscore en _db hace que st.cache_data no intente hashear el objeto DB.
@st.cache_data(ttl=30, show_spinner=False)
def q_obtener_viajes(_db, fecha_ini=None, fecha_fin=None, placa=None,
                     conductor=None, cliente=None, estado=None,
                     manifiesto=None, numero_factura=None):
    return _db.obtener_viajes(fecha_ini, fecha_fin, placa, conductor,
                               cliente, estado, manifiesto, numero_factura)

@st.cache_data(ttl=30, show_spinner=False)
def q_stats_dashboard(_db, fecha_ini, fecha_fin):
    return _db.stats_dashboard(fecha_ini, fecha_fin)

@st.cache_data(ttl=30, show_spinner=False)
def q_obtener_tanqueos(_db, placa=None, fecha_ini=None, fecha_fin=None):
    return _db.obtener_tanqueos(placa, fecha_ini, fecha_fin)

@st.cache_data(ttl=30, show_spinner=False)
def q_saldo_combustible_placa(_db, placa: str):
    return _db.saldo_combustible_placa(placa)


def limpiar_cache():
    """
    Limpia solo las funciones de consulta, no el pool de conexiones.
    Se llama después de cualquier escritura para refrescar datos.
    """
    q_obtener_viajes.clear()
    q_stats_dashboard.clear()
    q_obtener_tanqueos.clear()
    q_saldo_combustible_placa.clear()


def hora_a_time(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return None
    if isinstance(val, time): return val
    try:
        s = str(val)[:5]; h, m = s.split(":"); return time(int(h), int(m))
    except: return None

def str_hora(val):
    t = hora_a_time(val)
    return t.strftime("%H:%M") if t else "—"

def calcular_duracion(h_ini, h_fin, dias_extra=0):
    t1 = hora_a_time(h_ini)
    t2 = hora_a_time(h_fin)
    if not t1 or not t2: return None
    d1 = timedelta(hours=t1.hour, minutes=t1.minute)
    d2 = timedelta(hours=t2.hour, minutes=t2.minute)
    diff = d2 - d1 + timedelta(days=int(dias_extra or 0))
    if diff.total_seconds() < 0 and (dias_extra or 0) == 0:
        diff += timedelta(days=1)
    return int(diff.total_seconds() / 60)

def mins_a_str(mins):
    if mins is None: return "—"
    mins = int(mins)
    dias = mins // (60 * 24)
    resto = mins % (60 * 24)
    h, m = divmod(resto, 60)
    if dias > 0:
        return f"{dias}d {h}h {m:02d}m"
    return f"{h}h {m:02d}m"

# ==================== WIDGET HORAS ====================
def widget_horas(prefix, val_cita=None, val_sal_cargue=None,
                 val_ll_desc=None, val_sal_desc=None,
                 dias_sc=0, dias_ld=0, dias_sd=0):
    st.markdown("#### ⏱️ Tiempos de Operación")
    st.caption("💡 Si el vehículo tardó más de un día entre etapas, indica los +días extra.")

    h1, h2, h3, h4 = st.columns(4)
    with h1: hora_cita = st.time_input("Cita Cargue",        value=val_cita,        step=300, key=f"{prefix}_hcc")
    with h2: hora_sc   = st.time_input("Salida Cargue",      value=val_sal_cargue,  step=300, key=f"{prefix}_hsc")
    with h3: hora_ld   = st.time_input("Llegada Descargue",  value=val_ll_desc,     step=300, key=f"{prefix}_hld")
    with h4: hora_sd   = st.time_input("Salida Descargue",   value=val_sal_desc,    step=300, key=f"{prefix}_hsd")

    _, d2, d3, d4 = st.columns(4)
    with d2: d_sc = st.number_input("➕ Días extra Salida Cargue",     min_value=0, max_value=30, value=int(dias_sc or 0), step=1, key=f"{prefix}_dsc")
    with d3: d_ld = st.number_input("➕ Días extra Llegada Descargue", min_value=0, max_value=30, value=int(dias_ld or 0), step=1, key=f"{prefix}_dld")
    with d4: d_sd = st.number_input("➕ Días extra Salida Descargue",  min_value=0, max_value=30, value=int(dias_sd or 0), step=1, key=f"{prefix}_dsd")

    if hora_cita or hora_sc or hora_ld or hora_sd:
        t_espera    = calcular_duracion(hora_cita, hora_sc, d_sc)
        t_transito  = calcular_duracion(hora_sc,   hora_ld, d_ld)
        t_descargue = calcular_duracion(hora_ld,   hora_sd, d_sd)
        t_total     = (t_espera + t_transito + t_descargue) if (t_espera and t_transito and t_descargue) else None

        cols_prev = st.columns(4)
        previews  = [("⏳ Espera Cargue", t_espera), ("🚛 Tránsito", t_transito),
                     ("📦 Descargue", t_descargue), ("🕐 Total", t_total)]
        for col, (lbl, val) in zip(cols_prev, previews):
            color = "#e94560" if val is not None else "#4a4a6a"
            col.markdown(
                f"<div style='text-align:center;padding:6px;background:#1a1a2e;border-radius:8px;"
                f"border-top:3px solid {color};margin-top:4px;'>"
                f"<div style='font-size:0.68rem;color:#8892b0;'>{lbl}</div>"
                f"<div style='font-size:1rem;font-weight:700;color:{color};font-family:Rajdhani,sans-serif;'>{mins_a_str(val)}</div>"
                f"</div>",
                unsafe_allow_html=True
            )

    return hora_cita, hora_sc, hora_ld, hora_sd, d_sc, d_ld, d_sd

# ==================== WIDGET LIQUIDACIÓN ====================
def widget_liquidacion(flete, comision, bono, combustible_pesos, combustible_galones=0):
    utilidad   = int(flete or 0) - int(comision or 0)
    color_util = "#2ecc71" if utilidad >= 0 else "#e74c3c"
    show_bono  = "flex" if int(bono or 0) > 0 else "none"
    show_comb  = "flex" if (int(combustible_pesos or 0) > 0 or float(combustible_galones or 0) > 0) else "none"
    gal_txt    = f"{float(combustible_galones or 0):.2f} gal" if float(combustible_galones or 0) > 0 else ""
    gal_sep    = " &middot; " if gal_txt else ""

    html = (
        '<div class="liquidacion-box">'
        '<div style="font-family:Rajdhani,sans-serif;font-size:0.75rem;color:#e94560;letter-spacing:2px;margin-bottom:8px;">LIQUIDACION DEL VIAJE</div>'
        '<div class="liq-row"><span>&#128666; Flete</span>'
        f'<span style="color:#2ecc71;">{fmt_moneda(flete)}</span></div>'
        '<div class="liq-row"><span>&#128100; Comision conductor</span>'
        f'<span style="color:#e74c3c;">- {fmt_moneda(comision)}</span></div>'
        '<div class="liq-total"><span>&#128176; Utilidad (Flete - Comision)</span>'
        f'<span style="color:{color_util};">{fmt_moneda(utilidad)}</span></div>'
        f'<div class="liq-row" style="display:{show_bono};">'
        f'<span>&#127769; Bono transporte <span style="font-size:0.72rem;color:#8892b0;">(informativo)</span></span>'
        f'<span style="color:#f39c12;">{fmt_moneda(bono)}</span></div>'
        f'<div class="liq-row" style="display:{show_comb};">'
        f'<span>&#9981; Combustible{gal_sep}{gal_txt} <span style="font-size:0.72rem;color:#8892b0;">(informativo)</span></span>'
        f'<span style="color:#8892b0;">{fmt_moneda(combustible_pesos)}</span></div>'
        '</div>'
    )
    st.markdown(html, unsafe_allow_html=True)

# ==================== EXCEL ====================
def generar_excel(df: pd.DataFrame, titulo: str = "JP Transportamos") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Viajes"

    ft_titulo  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ft_header  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ft_normal  = Font(name="Calibri", size=9)
    ft_total   = Font(name="Calibri", bold=True, size=10)
    ft_anulado = Font(name="Calibri", size=9, color="C0392B")
    ft_incump  = Font(name="Calibri", size=9, color="D35400")

    fill_titulo  = PatternFill("solid", start_color="0A0A1A")
    fill_header  = PatternFill("solid", start_color="1A1A2E")
    fill_alt     = PatternFill("solid", start_color="EBF5FB")
    fill_total   = PatternFill("solid", start_color="D5DBDB")
    fill_anulado = PatternFill("solid", start_color="FADBD8")
    fill_incump  = PatternFill("solid", start_color="FDEBD0")
    fill_dinero  = PatternFill("solid", start_color="E8F8F5")

    borde  = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    now_col = datetime.now(pytz.timezone("America/Bogota"))

    columnas = [
        ("fecha","FECHA",12),("placa","PLACA",12),("conductor","CONDUCTOR",20),
        ("cliente","CLIENTE",22),("origen","ORIGEN",18),("destino","DESTINO",18),
        ("hora_cita_cargue","H.CITA",12),("hora_salida_cargue","H.SAL.CARGUE",13),
        ("hora_llegada_descargue","H.LLEGADA",13),("hora_salida_descargue","H.SAL.DESC",13),
        ("contenedor","CONTENEDOR",16),("carga","CARGA",12),
        ("numero_factura","Nº FACTURA",16),("manifiesto","MANIFIESTO",12),
        ("observacion","OBSERVACIÓN",25),("estado","ESTADO",13),
    ]
    cols_dinero  = ["FLETE","COMISIÓN","BONO TRANSP.","COMB. ($)","COMB. (GAL)","UTILIDAD NETA"]
    cols_tiempos = ["ESPERA CARGUE","TRÁNSITO","DESCARGUE","TOTAL OPERACIÓN"]

    total_cols = len(columnas) + len(cols_dinero) + len(cols_tiempos)
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"🚛 {titulo}   |   Generado: {now_col.strftime('%d/%m/%Y %H:%M')} (COL)   |   Total: {len(df)} viajes"
    ws["A1"].font = ft_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 30

    for idx, (key, nombre, ancho) in enumerate(columnas, start=1):
        c = ws.cell(row=2, column=idx, value=nombre)
        c.font = ft_header; c.fill = fill_header
        c.alignment = centro; c.border = borde
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    col_d_start = len(columnas) + 1
    for i, nombre in enumerate(cols_dinero, start=col_d_start):
        c = ws.cell(row=2, column=i, value=nombre)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="1E8449")
        c.alignment = centro; c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = 14

    col_t_start = col_d_start + len(cols_dinero)
    for i, nombre in enumerate(cols_tiempos, start=col_t_start):
        c = ws.cell(row=2, column=i, value=nombre)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="1A5276")
        c.alignment = centro; c.border = borde
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.row_dimensions[2].height = 28

    tot_flete = tot_comision = tot_bono = tot_comb_pesos = tot_comb_gal = 0

    for row_idx, (_, fila) in enumerate(df.iterrows(), start=3):
        estado_val = str(fila.get("estado", ""))
        es_an = "Anulado"    in estado_val
        es_in = "Incumplido" in estado_val
        fill_f = fill_anulado if es_an else (fill_incump if es_in else (fill_alt if row_idx % 2 == 0 else None))

        for col_idx, (key, _, _) in enumerate(columnas, start=1):
            val = fila.get(key, "")
            if not isinstance(val, str) and pd.isna(val): val = ""
            if key.startswith("hora_") and val:
                try: val = str(val)[:5]
                except: val = ""
            c = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            c.border = borde
            c.alignment = centro if key in ("fecha","placa","estado") or key.startswith("hora_") else izq
            c.font = ft_anulado if es_an else (ft_incump if es_in else ft_normal)
            if fill_f: c.fill = fill_f

        flete_v      = int(fila.get("flete", 0) or 0)
        comision_v   = int(fila.get("comision", 0) or 0)
        bono_v       = int(fila.get("bono_transporte", 0) or 0)
        comb_pesos_v = int(fila.get("combustible_pesos", 0) or 0)
        comb_gal_v   = float(fila.get("combustible_galones", 0) or 0)
        utilidad_v   = flete_v - comision_v
        tot_flete      += flete_v
        tot_comision   += comision_v
        tot_bono       += bono_v
        tot_comb_pesos += comb_pesos_v
        tot_comb_gal   += comb_gal_v

        fill_d = fill_dinero if fill_f is None else fill_f
        for ci, (val, fmt, es_neg) in enumerate([
            (flete_v,    '#,##0', False),
            (comision_v, '#,##0', True),
            (bono_v,     '#,##0', True),
            (comb_pesos_v,'#,##0',True),
            (comb_gal_v, '0.00', False),
            (utilidad_v, '#,##0', utilidad_v < 0),
        ], start=col_d_start):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = Font(name="Calibri", size=9,
                          color="C0392B" if es_neg and val < 0 else ("1E8449" if ci == col_d_start + 5 else "000000"))
            c.border = borde; c.alignment = centro
            c.number_format = fmt
            if fill_d: c.fill = fill_d

        d_sc_ = int(fila.get("dias_salida_cargue", 0) or 0)
        d_ld_ = int(fila.get("dias_llegada_descargue", 0) or 0)
        d_sd_ = int(fila.get("dias_salida_descargue", 0) or 0)
        t_e = calcular_duracion(fila.get("hora_cita_cargue"),       fila.get("hora_salida_cargue"),     d_sc_)
        t_t = calcular_duracion(fila.get("hora_salida_cargue"),     fila.get("hora_llegada_descargue"), d_ld_)
        t_d = calcular_duracion(fila.get("hora_llegada_descargue"), fila.get("hora_salida_descargue"),  d_sd_)
        t_tot = (t_e + t_t + t_d) if (t_e and t_t and t_d) else None

        for ci, val in enumerate([mins_a_str(t_e), mins_a_str(t_t), mins_a_str(t_d), mins_a_str(t_tot)], start=col_t_start):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = ft_normal; c.border = borde; c.alignment = centro

        ws.row_dimensions[row_idx].height = 18

    total_row = len(df) + 3
    tot_utilidad = tot_flete - tot_comision
    try:
        ws.merge_cells(f"A{total_row}:{get_column_letter(len(columnas))}{total_row}")
    except: pass
    ct = ws.cell(total_row, 1,
        f"TOTAL: {len(df)} viajes  |  ✅ {len(df[df['estado'].str.contains('Completado',na=False)])}  ❌ {len(df[df['estado'].str.contains('Anulado',na=False)])}  ⚠️ {len(df[df['estado'].str.contains('Incumplido',na=False)])}")
    ct.font = ft_total; ct.fill = fill_total; ct.alignment = centro

    for ci, (val, fmt_n) in enumerate([
        (tot_flete,'#,##0'),(tot_comision,'#,##0'),(tot_bono,'#,##0'),
        (tot_comb_pesos,'#,##0'),(tot_comb_gal,'0.00'),(tot_utilidad,'#,##0')
    ], start=col_d_start):
        c = ws.cell(total_row, ci, val)
        c.font = Font(name="Calibri", bold=True, size=10,
                      color="1E8449" if ci == col_d_start + 5 and val >= 0 else ("C0392B" if val < 0 else "000000"))
        c.fill = fill_total; c.border = borde; c.alignment = centro
        c.number_format = fmt_n

    # ==================== HOJA CONDUCTORES ====================
    ws_c = wb.create_sheet("Conductores")
    ws_c["A1"] = "Trazabilidad por Conductor"
    ws_c["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws_c["A1"].fill = PatternFill("solid", start_color="0A0A1A")
    ws_c["A1"].alignment = centro
    ws_c.row_dimensions[1].height = 26

    hdrs_c = ["CONDUCTOR","VIAJES","COMPLET.","ANULADOS","INCUMPL.","% CUMPL.",
              "TOT. FLETE","TOT. COMISIÓN","TOT. BONO","COMB. ($)","COMB. (GAL)","UTILIDAD NETA"]
    for ci, h in enumerate(hdrs_c, start=1):
        c = ws_c.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="1A1A2E")
        c.alignment = centro; c.border = borde
    ws_c.row_dimensions[2].height = 20

    if "conductor" in df.columns:
        df_cond = df.groupby("conductor").agg(
            viajes=("conductor","count"),
            comp=("estado",   lambda x: x.str.contains("Completado", na=False).sum()),
            anul=("estado",   lambda x: x.str.contains("Anulado",    na=False).sum()),
            incu=("estado",   lambda x: x.str.contains("Incumplido", na=False).sum()),
            t_flete=("flete", "sum"),
            t_com=("comision","sum"),
            t_bono=("bono_transporte","sum"),
            t_comb_p=("combustible_pesos","sum"),
            t_comb_g=("combustible_galones","sum"),
        ).reset_index().sort_values("viajes", ascending=False)
        df_cond["utilidad"] = df_cond["t_flete"] - df_cond["t_com"]
        df_cond["pct"] = df_cond.apply(lambda r: f"{round(r.comp/r.viajes*100,1)}%" if r.viajes > 0 else "0%", axis=1)

        for i, r in enumerate(df_cond.itertuples(), start=3):
            fill_ci = PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None
            vals = [r.conductor, r.viajes, r.comp, r.anul, r.incu, r.pct,
                    int(r.t_flete), int(r.t_com), int(r.t_bono),
                    int(r.t_comb_p), round(float(r.t_comb_g),2), int(r.utilidad)]
            for ci2, v in enumerate(vals, start=1):
                c = ws_c.cell(i, ci2, v)
                c.font = ft_normal; c.border = borde
                c.alignment = izq if ci2 == 1 else centro
                if ci2 in (7,8,9,10,12): c.number_format = '#,##0'
                if ci2 == 11: c.number_format = '0.00'
                if fill_ci: c.fill = fill_ci

    anchos_c = {"A":24,"B":8,"C":10,"D":10,"E":10,"F":9,
                "G":14,"H":15,"I":13,"J":14,"K":12,"L":14}
    for col_l, w in anchos_c.items():
        ws_c.column_dimensions[col_l].width = w
    ws_c.freeze_panes = "A3"

    # ==================== HOJA VEHÍCULOS ====================
    ws_v = wb.create_sheet("Vehículos")
    ws_v["A1"] = "Trazabilidad por Vehículo"
    ws_v["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws_v["A1"].fill = PatternFill("solid", start_color="0A0A1A")
    ws_v["A1"].alignment = centro
    ws_v.row_dimensions[1].height = 26

    hdrs_v = ["PLACA","VIAJES","COMPLET.","TOT. FLETE","COMB. ($)","COMB. (GAL)","UTILIDAD","CONDUCTORES"]
    for ci, h in enumerate(hdrs_v, start=1):
        c = ws_v.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="1A5276")
        c.alignment = centro; c.border = borde
    ws_v.row_dimensions[2].height = 20

    if "placa" in df.columns:
        df_veh = df.groupby("placa").agg(
            viajes=("placa","count"),
            comp=("estado", lambda x: x.str.contains("Completado", na=False).sum()),
            t_flete=("flete","sum"),
            t_com=("comision","sum"),
            t_comb_p=("combustible_pesos","sum"),
            t_comb_g=("combustible_galones","sum"),
            conductores=("conductor", lambda x: ", ".join(x.dropna().unique()[:3])),
        ).reset_index().sort_values("viajes", ascending=False)
        df_veh["utilidad"] = df_veh["t_flete"] - df_veh["t_com"]

        for i, r in enumerate(df_veh.itertuples(), start=3):
            fill_vi = PatternFill("solid", start_color="D6EAF8") if i % 2 == 0 else None
            vals = [r.placa, r.viajes, r.comp, int(r.t_flete),
                    int(r.t_comb_p), round(float(r.t_comb_g),2), int(r.utilidad), r.conductores]
            for ci2, v in enumerate(vals, start=1):
                c = ws_v.cell(i, ci2, v)
                c.font = ft_normal; c.border = borde
                c.alignment = izq if ci2 in (1,8) else centro
                if ci2 in (4,5,7): c.number_format = '#,##0'
                if ci2 == 6: c.number_format = '0.00'
                if fill_vi: c.fill = fill_vi

    for col_l, w in zip(["A","B","C","D","E","F","G","H"],[12,8,10,14,14,12,14,30]):
        ws_v.column_dimensions[col_l].width = w
    ws_v.freeze_panes = "A3"

    # ==================== HOJA FINANCIERO ====================
    ws_f = wb.create_sheet("Financiero")
    ws_f["A1"] = "Resumen Financiero"
    ws_f["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws_f["A1"].fill = PatternFill("solid", start_color="0A0A1A")
    ws_f["A1"].alignment = centro
    ws_f.row_dimensions[1].height = 26

    kpis_f = [
        ("Total Fletes", tot_flete),
        ("Total Comisiones Conductores", tot_comision),
        ("Utilidad Neta (Flete − Comisión)", tot_utilidad),
        ("Margen % (Utilidad / Flete)", f"{round(tot_utilidad/tot_flete*100,1)}%" if tot_flete > 0 else "0%"),
        ("Promedio Flete por Viaje", int(tot_flete/len(df)) if len(df) > 0 else 0),
        ("Promedio Utilidad por Viaje", int(tot_utilidad/len(df)) if len(df) > 0 else 0),
        ("", ""),
        ("── Datos Informativos ──", ""),
        ("Total Bonos Transporte", tot_bono),
        ("Total Combustible en Pesos", tot_comb_pesos),
        ("Total Combustible en Galones", round(tot_comb_gal, 2)),
    ]
    ws_f.cell(2,1,"CONCEPTO").font = ft_header
    ws_f.cell(2,1).fill = PatternFill("solid", start_color="1E8449")
    ws_f.cell(2,1).alignment = centro; ws_f.cell(2,1).border = borde
    ws_f.cell(2,2,"VALOR").font = ft_header
    ws_f.cell(2,2).fill = PatternFill("solid", start_color="1E8449")
    ws_f.cell(2,2).alignment = centro; ws_f.cell(2,2).border = borde
    ws_f.row_dimensions[2].height = 20

    for i, (lbl, val) in enumerate(kpis_f, start=3):
        fill_fi = PatternFill("solid", start_color="E8F8F5") if i % 2 == 0 else None
        c1 = ws_f.cell(i,1,lbl); c2 = ws_f.cell(i,2,val)
        c1.font = ft_normal; c2.font = ft_total
        c1.border = borde; c2.border = borde
        c1.alignment = izq; c2.alignment = centro
        if isinstance(val, int): c2.number_format = '#,##0'
        if fill_fi: c1.fill = fill_fi; c2.fill = fill_fi

    for col_l, w in zip(["A","B"],[32,16]):
        ws_f.column_dimensions[col_l].width = w

    # ==================== HOJA TIEMPOS ====================
    ws_t = wb.create_sheet("Tiempos")
    ws_t["A1"] = "Análisis de Tiempos por Viaje"
    ws_t["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws_t["A1"].fill = PatternFill("solid", start_color="0A0A1A")
    ws_t["A1"].alignment = centro
    ws_t.row_dimensions[1].height = 26
    hdrs_t = ["FECHA","PLACA","CONDUCTOR","CLIENTE","ESPERA CARGUE","TRÁNSITO","DESCARGUE","TOTAL OPERACIÓN"]
    for ci, h in enumerate(hdrs_t, start=1):
        c = ws_t.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="203A43")
        c.alignment = centro; c.border = borde

    for i, (_, row) in enumerate(df.iterrows(), start=3):
        d_sc_ = int(row.get("dias_salida_cargue",0) or 0)
        d_ld_ = int(row.get("dias_llegada_descargue",0) or 0)
        d_sd_ = int(row.get("dias_salida_descargue",0) or 0)
        t_e = calcular_duracion(row.get("hora_cita_cargue"),       row.get("hora_salida_cargue"),     d_sc_)
        t_t = calcular_duracion(row.get("hora_salida_cargue"),     row.get("hora_llegada_descargue"), d_ld_)
        t_d = calcular_duracion(row.get("hora_llegada_descargue"), row.get("hora_salida_descargue"),  d_sd_)
        t_tot = (t_e + t_t + t_d) if (t_e and t_t and t_d) else None
        fill_ti = PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None
        vals = [str(row.get("fecha","")), str(row.get("placa","")),
                str(row.get("conductor","")), str(row.get("cliente","")),
                mins_a_str(t_e), mins_a_str(t_t), mins_a_str(t_d), mins_a_str(t_tot)]
        for ci2, v in enumerate(vals, start=1):
            c = ws_t.cell(i, ci2, v)
            c.font = ft_normal; c.border = borde
            c.alignment = izq if ci2 <= 4 else centro
            if fill_ti: c.fill = fill_ti
    for col_l, w in zip(["A","B","C","D","E","F","G","H"],[12,10,22,20,15,15,15,16]):
        ws_t.column_dimensions[col_l].width = w
    ws_t.freeze_panes = "A3"

    ws.freeze_panes = "A3"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==================== MAIN ====================
def main():
    st.markdown("""
    <div class="main-header">
        <div class="empresa-badge">EMPRESA</div>
        <h1>🚛 JP Transportamos</h1>
        <p class="subtitle">Control de Viajes · Comisiones · Trazabilidad · Combustible</p>
    </div>
    """, unsafe_allow_html=True)

    if "editando_id" not in st.session_state:
        st.session_state.editando_id = None

    db = get_db()

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📝 Nuevo Viaje",
        "🔍 Historial y Reportes",
        "📊 Dashboard",
        "🧾 Trazabilidad",
        "⛽ Control Combustible"
    ])

    # ===================== TAB 1: NUEVO VIAJE =====================
    with tab1:
        st.markdown("### Registrar Nuevo Viaje")

        f1, f2, f3, f4 = st.columns(4)
        with f1: fecha_pre = st.date_input("📅 Fecha", datetime.now(), key="pre_fecha")
        with f2:
            placas_lista = list(PLACA_CONDUCTOR.keys())
            placa_pre = st.selectbox("🚛 Placa", placas_lista, key="pre_placa")
        with f3:
            conductor_fijo = PLACA_CONDUCTOR.get(placa_pre)
            cond_opts = ["— Seleccionar —"] + TODOS_CONDUCTORES
            cond_default = cond_opts.index(conductor_fijo) if conductor_fijo in cond_opts else 0
            conductor_sel = st.selectbox("👤 Conductor", cond_opts, index=cond_default, key="pre_conductor")
        with f4:
            cli_sel = st.selectbox("🏢 Cliente", CLIENTES_FRECUENTES + [LABEL_MANUAL_CLI], key="pre_cliente")

        cliente_pre = st.text_input("✏️ Escribir cliente", placeholder="Nombre del cliente...", key="pre_cli_manual") \
            if cli_sel == LABEL_MANUAL_CLI else cli_sel

        st.markdown("#### 🗺️ Ruta")
        ruta_opts = [f"{o}  →  {d}" for o, d in RUTAS_FRECUENTES] + [LABEL_MANUAL]
        ruta_sel = st.selectbox("Ruta frecuente", ruta_opts, index=len(ruta_opts)-1, key="pre_ruta")
        c5, c6 = st.columns(2)
        if ruta_sel == LABEL_MANUAL:
            with c5: origen_pre  = st.text_input("📍 Origen",  placeholder="Escribe el origen...",  key="pre_origen")
            with c6: destino_pre = st.text_input("🏁 Destino", placeholder="Escribe el destino...", key="pre_destino")
        else:
            _o, _d = ruta_sel.split("  →  ")
            with c5: st.info(f"📍 **Origen:** {_o}")
            with c6: st.info(f"🏁 **Destino:** {_d}")
            origen_pre, destino_pre = _o, _d

        hora_cita, hora_sc, hora_ld, hora_sd, d_sc, d_ld, d_sd = widget_horas(prefix="new")

        bono_auto = calcular_bono_transporte(hora_cita)
        if bono_auto > 0:
            st.markdown(
                f"<div style='margin:6px 0;'>🌙 <b>Bono de transporte automático:</b> "
                f"<span class='bono-badge'>{fmt_moneda(bono_auto)}</span> "
                f"<span style='color:#8892b0;font-size:0.8rem;'>(hora de cita: {hora_cita.strftime('%H:%M') if hora_cita else '—'})</span></div>",
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                "<div style='margin:6px 0;'>🌙 Bono de transporte: <span class='bono-zero'>$0 — hora normal</span></div>",
                unsafe_allow_html=True
            )

        st.markdown("#### 💰 Datos Financieros")
        fin1, fin2, fin3, fin4 = st.columns(4)
        with fin1: flete_pre       = input_moneda("🚚 Flete ($)",              key="pre_flete")
        with fin2: comision_pre    = input_moneda("👤 Comisión conductor ($)", key="pre_comision")
        with fin3: comb_pesos_pre  = input_moneda("⛽ Combustible ($)",        key="pre_comb_pesos",
                                                   help_text="Costo del combustible en pesos (informativo)")
        with fin4:
            comb_gal_pre = st.number_input("⛽ Combustible (galones)", min_value=0.0,
                                            step=0.5, format="%.2f", key="pre_comb_gal",
                                            help="Galones consumidos en este viaje")

        widget_liquidacion(flete_pre, comision_pre, bono_auto, comb_pesos_pre, comb_gal_pre)

        with st.form("form_viaje", clear_on_submit=True):
            fecha      = fecha_pre
            placa      = placa_pre
            conductor  = "" if conductor_sel == "— Seleccionar —" else conductor_sel
            cliente    = cliente_pre
            origen     = origen_pre
            destino    = destino_pre

            st.markdown("#### 📦 Información de Carga")
            d1, d2, d3, d4 = st.columns(4)
            with d1: contenedor      = st.text_input("Contenedor")
            with d2: carga           = st.text_input("Carga")
            with d3: numero_factura  = st.text_input("Nº Factura")
            with d4: manifiesto      = st.text_input("Manifiesto")

            e1, e2 = st.columns([1, 3])
            with e1: estado      = st.selectbox("🚦 Estado", ESTADOS_VIAJE)
            with e2: observacion = st.text_area("📝 Observaciones", height=80)

            submitted = st.form_submit_button("💾 Guardar Viaje", type="primary", use_container_width=True)

        if submitted:
            if not placa:
                st.error("⚠️ La placa es obligatoria.")
            else:
                datos = {
                    "fecha": fecha, "placa": placa, "conductor": conductor,
                    "cliente": cliente, "origen": origen, "destino": destino,
                    "hora_cita_cargue": hora_cita,
                    "hora_salida_cargue": hora_sc,
                    "hora_llegada_descargue": hora_ld,
                    "hora_salida_descargue": hora_sd,
                    "contenedor": contenedor, "carga": carga,
                    "numero_factura": numero_factura,
                    "manifiesto": manifiesto, "observacion": observacion,
                    "estado": estado.split(" ", 1)[1] if " " in estado else estado,
                    "dias_salida_cargue": d_sc,
                    "dias_llegada_descargue": d_ld,
                    "dias_salida_descargue": d_sd,
                    "flete": flete_pre,
                    "comision": comision_pre,
                    "bono_transporte": bono_auto,
                    "combustible_pesos": comb_pesos_pre,
                    "combustible_galones": comb_gal_pre,
                }
                if db.guardar_viaje(datos):
                    limpiar_cache()
                    st.success(f"✅ Viaje guardado — {placa} | {conductor} | {origen} → {destino} | Flete: {fmt_moneda(flete_pre)}")
                    st.balloons()

    # ===================== TAB 2: HISTORIAL =====================
    with tab2:
        st.markdown("### 🔍 Historial de Viajes")

        with st.expander("🛠️ Filtros", expanded=True):
            f1, f2, f3, f4 = st.columns(4)
            with f1: fi   = st.date_input("Desde", datetime.now() - timedelta(days=30), key="h_fi")
            with f2: ff   = st.date_input("Hasta", datetime.now(), key="h_ff")
            with f3:
                placas_h = ["Todas"] + list(PLACA_CONDUCTOR.keys())
                fp = st.selectbox("Placa", placas_h, key="h_fp")
            with f4: fc   = st.text_input("Conductor", key="h_fc")

            f5, f6, f7, f8 = st.columns(4)
            with f5: fcli = st.text_input("Cliente",   key="h_fcli")
            with f6:
                estados_f = ["Todos"] + [e.split(" ", 1)[1] for e in ESTADOS_VIAJE]
                fe = st.selectbox("Estado", estados_f, key="h_fe")
            with f7: f_manifiesto = st.text_input("📋 Manifiesto", key="h_manifiesto", placeholder="Buscar por manifiesto...")
            with f8: f_factura    = st.text_input("🧾 Nº Factura",  key="h_factura",    placeholder="Buscar por factura...")

        df = q_obtener_viajes(db, fi, ff, fp, fc, fcli,
                               fe if fe != "Todos" else None,
                               f_manifiesto or None,
                               f_factura or None)

        if not df.empty:
            k1, k2, k3, k4, k5, k6 = st.columns(6)
            k1.metric("Total Viajes",    len(df))
            k2.metric("✅ Completados",  len(df[df["estado"].str.contains("Completado", na=False)]))
            k3.metric("❌ Anulados",     len(df[df["estado"].str.contains("Anulado",    na=False)]))
            k4.metric("⚠️ Incumplidos", len(df[df["estado"].str.contains("Incumplido", na=False)]))
            k5.metric("💰 Total Fletes", fmt_moneda(df["flete"].sum()))
            k6.metric("📈 Utilidad",     fmt_moneda(df["flete"].sum() - df["comision"].sum()))

            st.divider()
            col_exp1, col_exp2 = st.columns([2, 5])
            with col_exp1:
                nombre_rep = st.text_input("Nombre del reporte", value="JP_Transportamos", key="rep_nombre")
            with col_exp2:
                st.markdown("<br>", unsafe_allow_html=True)
                excel_data = generar_excel(df, titulo=nombre_rep)
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=excel_data,
                    file_name=f"{nombre_rep}_{datetime.now(pytz.timezone('America/Bogota')).strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            st.divider()
            cols_tabla = ["id","fecha","placa","conductor","cliente","origen","destino",
                          "contenedor","carga","numero_factura","manifiesto","estado",
                          "flete","comision","bono_transporte","combustible_pesos","combustible_galones"]
            cols_ex = [c for c in cols_tabla if c in df.columns]
            st.dataframe(df[cols_ex], use_container_width=True, hide_index=True)

            st.divider()
            st.subheader("✏️ Ver Detalle / Editar")
            df["_label"] = df.apply(
                lambda r: f"ID {r['id']} | {r['fecha']} | {r['placa']} | {r.get('cliente','')} | {r.get('origen','')} → {r.get('destino','')} | {r.get('estado','')}",
                axis=1
            )
            sel = st.selectbox("Seleccionar viaje:", df["_label"].tolist(), key="h_sel")

            if sel:
                vid = int(sel.split(" | ")[0].replace("ID ", ""))
                row = df[df["id"] == vid].iloc[0]
                editando = st.session_state.editando_id == vid

                if not editando:
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.info(f"**Placa:** {row['placa']}")
                        st.write(f"**Conductor:** {row.get('conductor','')}")
                        st.write(f"**Cliente:** {row.get('cliente','')}")
                        st.write(f"**Fecha:** {row['fecha']}")
                    with c2:
                        st.write(f"**Origen:** {row.get('origen','')}")
                        st.write(f"**Destino:** {row.get('destino','')}")
                        st.write(f"**Contenedor:** {row.get('contenedor','')}")
                        st.write(f"**Nº Factura:** {row.get('numero_factura','')}")
                        st.write(f"**Manifiesto:** {row.get('manifiesto','')}")
                    with c3:
                        estado_raw = str(row.get('estado',''))
                        color = "🟢" if "Completado" in estado_raw else ("🔴" if "Anulado" in estado_raw else "🟡")
                        st.write(f"**Estado:** {color} {estado_raw}")
                        st.write(f"**Observación:** {row.get('observacion','')}")

                    d_sc_r = int(row.get("dias_salida_cargue",0) or 0)
                    d_ld_r = int(row.get("dias_llegada_descargue",0) or 0)
                    d_sd_r = int(row.get("dias_salida_descargue",0) or 0)
                    t_e = calcular_duracion(row["hora_cita_cargue"],       row["hora_salida_cargue"],     d_sc_r)
                    t_t = calcular_duracion(row["hora_salida_cargue"],     row["hora_llegada_descargue"], d_ld_r)
                    t_d = calcular_duracion(row["hora_llegada_descargue"], row["hora_salida_descargue"],  d_sd_r)
                    t_tot = (t_e + t_t + t_d) if (t_e and t_t and t_d) else None
                    st.write(f"**Tiempos:** ⏳ `{mins_a_str(t_e)}` | 🚛 `{mins_a_str(t_t)}` | 📦 `{mins_a_str(t_d)}` | Total `{mins_a_str(t_tot)}`")

                    flete_r   = int(row.get("flete",0) or 0)
                    comision_r= int(row.get("comision",0) or 0)
                    bono_r    = int(row.get("bono_transporte",0) or 0)
                    comb_p_r  = int(row.get("combustible_pesos",0) or 0)
                    comb_g_r  = float(row.get("combustible_galones",0) or 0)
                    widget_liquidacion(flete_r, comision_r, bono_r, comb_p_r, comb_g_r)

                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("✏️ Editar", key=f"eb_{vid}"):
                            st.session_state.editando_id = vid; st.rerun()
                    with bc2:
                        if st.button("🗑️ Eliminar", key=f"del_{vid}"):
                            db.eliminar_viaje(vid); limpiar_cache(); st.success("Eliminado."); st.rerun()
                else:
                    st.markdown("#### ✏️ Editando viaje")
                    ec1, ec2, ec3, ec4 = st.columns(4)
                    with ec1: e_fecha = st.date_input("Fecha", value=row["fecha"], key=f"ef_{vid}")
                    with ec2:
                        placas_e = list(PLACA_CONDUCTOR.keys())
                        placa_idx = placas_e.index(row["placa"]) if row["placa"] in placas_e else 0
                        e_placa = st.selectbox("Placa", placas_e, index=placa_idx, key=f"ep_{vid}")
                    with ec3:
                        cond_fijo_e = PLACA_CONDUCTOR.get(e_placa)
                        cond_actual = str(row.get("conductor") or "")
                        cond_opts_e = ["— Seleccionar —"] + TODOS_CONDUCTORES
                        default_e = cond_opts_e.index(cond_fijo_e) if cond_fijo_e in cond_opts_e else (
                            cond_opts_e.index(cond_actual) if cond_actual in cond_opts_e else 0)
                        e_cond_sel = st.selectbox("👤 Conductor", cond_opts_e, index=default_e, key=f"ec_{vid}")
                        e_conductor = "" if e_cond_sel == "— Seleccionar —" else e_cond_sel
                    with ec4:
                        cli_actual = str(row.get("cliente") or "")
                        cli_opts = CLIENTES_FRECUENTES + [LABEL_MANUAL_CLI]
                        cli_idx = cli_opts.index(cli_actual) if cli_actual in cli_opts else len(cli_opts)-1
                        e_cli_sel = st.selectbox("Cliente", cli_opts, index=cli_idx, key=f"ecl_{vid}")
                        e_cliente = st.text_input("Cliente (manual)", value=cli_actual if cli_actual not in CLIENTES_FRECUENTES else "", key=f"ecl_m_{vid}") \
                            if e_cli_sel == LABEL_MANUAL_CLI else e_cli_sel

                    er1, er2 = st.columns(2)
                    with er1: e_origen  = st.text_input("Origen",  value=str(row.get("origen","")  or ""), key=f"eo_{vid}")
                    with er2: e_destino = st.text_input("Destino", value=str(row.get("destino","") or ""), key=f"ed_{vid}")

                    e_hcc, e_hsc, e_hld, e_hsd, e_dsc, e_dld, e_dsd = widget_horas(
                        prefix=f"edit_{vid}",
                        val_cita=hora_a_time(row["hora_cita_cargue"]),
                        val_sal_cargue=hora_a_time(row["hora_salida_cargue"]),
                        val_ll_desc=hora_a_time(row["hora_llegada_descargue"]),
                        val_sal_desc=hora_a_time(row["hora_salida_descargue"]),
                        dias_sc=int(row.get("dias_salida_cargue",0) or 0),
                        dias_ld=int(row.get("dias_llegada_descargue",0) or 0),
                        dias_sd=int(row.get("dias_salida_descargue",0) or 0),
                    )

                    bono_edit = calcular_bono_transporte(e_hcc)
                    st.markdown(f"🌙 Bono automático recalculado: **{fmt_moneda(bono_edit)}**")

                    st.markdown("#### 💰 Datos Financieros")
                    fe1, fe2, fe3, fe4 = st.columns(4)
                    with fe1: e_flete       = input_moneda("Flete ($)",          key=f"efl_{vid}",    value=int(row.get("flete",0) or 0))
                    with fe2: e_comision    = input_moneda("Comisión ($)",        key=f"eco2_{vid}",   value=int(row.get("comision",0) or 0))
                    with fe3: e_comb_pesos  = input_moneda("Combustible ($)",     key=f"ecomb_{vid}",  value=int(row.get("combustible_pesos",0) or 0))
                    with fe4:
                        e_comb_gal = st.number_input("Combustible (gal)", min_value=0.0, step=0.5, format="%.2f",
                                                      key=f"ecombg_{vid}",
                                                      value=float(row.get("combustible_galones",0) or 0))
                    widget_liquidacion(e_flete, e_comision, bono_edit, e_comb_pesos, e_comb_gal)

                    with st.form(f"edit_{vid}"):
                        ed1, ed2, ed3, ed4 = st.columns(4)
                        with ed1: e_cont = st.text_input("Contenedor", value=str(row.get("contenedor","") or ""), key=f"eco_{vid}")
                        with ed2: e_carga = st.text_input("Carga",     value=str(row.get("carga","") or ""),      key=f"eca_{vid}")
                        with ed3: e_fact  = st.text_input("Nº Factura",value=str(row.get("numero_factura","") or ""), key=f"efact_{vid}")
                        with ed4: e_man   = st.text_input("Manifiesto",value=str(row.get("manifiesto","") or ""), key=f"ema_{vid}")

                        estados_l = [e.split(" ", 1)[1] for e in ESTADOS_VIAJE]
                        est_actual = str(row.get("estado") or "Completado")
                        est_idx = estados_l.index(est_actual) if est_actual in estados_l else 0
                        ee1, ee2 = st.columns([1, 3])
                        with ee1: e_estado = st.selectbox("Estado", ESTADOS_VIAJE, index=est_idx, key=f"est_{vid}")
                        with ee2: e_obs    = st.text_area("Observaciones", value=str(row.get("observacion","") or ""), key=f"eob_{vid}", height=80)

                        sg1, sg2 = st.columns(2)
                        with sg1: guardar  = st.form_submit_button("💾 Guardar Cambios", type="primary")
                        with sg2: cancelar = st.form_submit_button("❌ Cancelar")

                    if guardar:
                        datos_edit = {
                            "fecha": e_fecha, "placa": e_placa, "conductor": e_conductor,
                            "cliente": e_cliente, "origen": e_origen, "destino": e_destino,
                            "hora_cita_cargue": e_hcc, "hora_salida_cargue": e_hsc,
                            "hora_llegada_descargue": e_hld, "hora_salida_descargue": e_hsd,
                            "contenedor": e_cont, "carga": e_carga,
                            "numero_factura": e_fact, "manifiesto": e_man,
                            "observacion": e_obs,
                            "estado": e_estado.split(" ", 1)[1] if " " in e_estado else e_estado,
                            "dias_salida_cargue": e_dsc, "dias_llegada_descargue": e_dld, "dias_salida_descargue": e_dsd,
                            "flete": e_flete, "comision": e_comision,
                            "bono_transporte": bono_edit,
                            "combustible_pesos": e_comb_pesos,
                            "combustible_galones": e_comb_gal,
                        }
                        if db.actualizar_viaje(vid, datos_edit):
                            limpiar_cache()
                            st.success("✅ Viaje actualizado.")
                            st.session_state.editando_id = None; st.rerun()
                    if cancelar:
                        st.session_state.editando_id = None; st.rerun()
        else:
            st.warning("No hay viajes con los filtros seleccionados.")

    # ===================== TAB 3: DASHBOARD =====================
    with tab3:
        st.markdown("### 📊 Dashboard de Operaciones")
        try:
            import plotly.express as px

            col_r1, _ = st.columns([2, 4])
            with col_r1:
                rango = st.date_input("Período", value=(datetime.now().replace(day=1), datetime.now()), key="dash_rango")

            if not (isinstance(rango, (list, tuple)) and len(rango) == 2):
                st.info("Selecciona un rango de fechas completo.")
                return

            df_s = q_stats_dashboard(db, rango[0], rango[1])
            if df_s.empty:
                st.info("No hay datos en este período.")
                return

            total = len(df_s)
            comp  = len(df_s[df_s["estado"].str.contains("Completado", na=False)])
            tot_flete_d = int(df_s["flete"].sum())
            tot_com_d   = int(df_s["comision"].sum())
            tot_comb_d  = int(df_s["combustible_pesos"].sum())
            tot_comb_gd = float(df_s["combustible_galones"].sum())
            utilidad_d  = tot_flete_d - tot_com_d
            pct = round(comp / total * 100) if total > 0 else 0

            k1, k2, k3, k4, k5, k6 = st.columns(6)
            k1.metric("🚚 Viajes",        total)
            k2.metric("✅ Completados",   comp, f"{pct}%")
            k3.metric("💰 Fletes",        fmt_moneda(tot_flete_d))
            k4.metric("👤 Comisiones",    fmt_moneda(tot_com_d))
            k5.metric("⛽ Combustible",   f"{round(tot_comb_gd,1)} gal / {fmt_moneda(tot_comb_d)}")
            k6.metric("📈 Utilidad",      fmt_moneda(utilidad_d))

            st.divider()
            g1, g2 = st.columns(2)
            with g1:
                st.markdown("#### Distribución por Estado")
                est_c = df_s["estado"].value_counts().reset_index()
                est_c.columns = ["estado","cantidad"]
                fig1 = px.pie(est_c, values="cantidad", names="estado", hole=0.45,
                              color_discrete_sequence=["#2ecc71","#e74c3c","#f39c12","#3498db"])
                fig1.update_layout(margin=dict(t=10,b=10), height=300)
                st.plotly_chart(fig1, use_container_width=True)

            with g2:
                st.markdown("#### Viajes por Día")
                df_dia = df_s.groupby("fecha").size().reset_index(name="viajes")
                fig2 = px.bar(df_dia, x="fecha", y="viajes",
                              color_discrete_sequence=["#e94560"], text="viajes")
                fig2.update_traces(textposition="outside")
                fig2.update_layout(margin=dict(t=10,b=10), height=300)
                st.plotly_chart(fig2, use_container_width=True)

            st.divider()
            g3, g4 = st.columns(2)
            with g3:
                st.markdown("#### 💰 Fletes por Conductor")
                if "conductor" in df_s.columns:
                    df_fc = df_s.groupby("conductor")["flete"].sum().reset_index().sort_values("flete")
                    fig3 = px.bar(df_fc, x="flete", y="conductor", orientation="h",
                                  color="flete", color_continuous_scale="Reds", text="flete")
                    fig3.update_traces(texttemplate='$%{text:,.0f}', textposition="outside")
                    fig3.update_layout(margin=dict(t=10,b=10), height=max(250,len(df_fc)*40),
                                       coloraxis_showscale=False)
                    st.plotly_chart(fig3, use_container_width=True)

            with g4:
                st.markdown("#### 📈 Utilidad por Conductor")
                if "conductor" in df_s.columns:
                    df_uc = df_s.groupby("conductor").agg(
                        flete=("flete","sum"), comision=("comision","sum")
                    ).reset_index()
                    df_uc["utilidad"] = df_uc["flete"] - df_uc["comision"]
                    df_uc = df_uc.sort_values("utilidad")
                    fig4 = px.bar(df_uc, x="utilidad", y="conductor", orientation="h",
                                  color="utilidad", color_continuous_scale="Greens", text="utilidad")
                    fig4.update_traces(texttemplate='$%{text:,.0f}', textposition="outside")
                    fig4.update_layout(margin=dict(t=10,b=10), height=max(250,len(df_uc)*40),
                                       coloraxis_showscale=False)
                    st.plotly_chart(fig4, use_container_width=True)

            st.divider()
            g5, g6 = st.columns(2)
            with g5:
                st.markdown("#### ⛽ Galones Consumidos por Vehículo")
                df_vf = df_s.groupby("placa").agg(
                    flete=("flete","sum"),
                    gal=("combustible_galones","sum")
                ).reset_index()
                fig5 = px.bar(df_vf, x="placa", y=["flete","gal"],
                              barmode="group",
                              color_discrete_map={"flete":"#2ecc71","gal":"#e74c3c"})
                fig5.update_layout(margin=dict(t=10,b=10), height=300)
                st.plotly_chart(fig5, use_container_width=True)

            with g6:
                st.markdown("#### 🌙 Bonos de Transporte")
                df_bon = df_s[df_s["bono_transporte"] > 0].groupby("conductor")["bono_transporte"].sum().reset_index().sort_values("bono_transporte", ascending=False)
                if not df_bon.empty:
                    fig6 = px.bar(df_bon, x="conductor", y="bono_transporte",
                                  color_discrete_sequence=["#f39c12"], text="bono_transporte")
                    fig6.update_traces(texttemplate='$%{text:,.0f}', textposition="outside")
                    fig6.update_layout(margin=dict(t=10,b=10), height=300)
                    st.plotly_chart(fig6, use_container_width=True)
                else:
                    st.info("Sin bonos en este período.")

        except ImportError:
            st.warning("Instala plotly: `pip install plotly`")
        except Exception as e:
            st.error(f"Error en dashboard: {e}")

    # ===================== TAB 4: TRAZABILIDAD =====================
    with tab4:
        st.markdown("### 🧾 Trazabilidad")

        traza_tab1, traza_tab2 = st.tabs(["👤 Por Conductor", "🚛 Por Vehículo"])

        with traza_tab1:
            conductor_sel_t = st.selectbox("Seleccionar conductor:", TODOS_CONDUCTORES, key="t_conductor")

            with st.expander("📅 Filtro de fechas", expanded=False):
                tc1, tc2 = st.columns(2)
                with tc1: t_fi = st.date_input("Desde", datetime.now() - timedelta(days=90), key="t_fi")
                with tc2: t_ff = st.date_input("Hasta", datetime.now(), key="t_ff")

            df_cond_t = q_obtener_viajes(db, fecha_ini=t_fi, fecha_fin=t_ff, conductor=conductor_sel_t)

            if not df_cond_t.empty:
                total_c    = len(df_cond_t)
                comp_c     = len(df_cond_t[df_cond_t["estado"].str.contains("Completado", na=False)])
                flete_c    = int(df_cond_t["flete"].sum())
                comision_c = int(df_cond_t["comision"].sum())
                bono_c     = int(df_cond_t["bono_transporte"].sum())
                comb_p_c   = int(df_cond_t["combustible_pesos"].sum())
                comb_g_c   = float(df_cond_t["combustible_galones"].sum())
                utilidad_c = flete_c - comision_c

                st.markdown(f"""
                <div class="trazabilidad-header">
                    <h4>👤 {conductor_sel_t} — Resumen del período</h4>
                </div>
                """, unsafe_allow_html=True)

                r1c1, r1c2, r1c3, r1c4 = st.columns(4)
                r1c1.metric("🚚 Viajes Totales",    total_c)
                r1c2.metric("✅ Completados",        comp_c, f"{round(comp_c/total_c*100)}%" if total_c > 0 else "")
                r1c3.metric("💰 Total Fletes",       fmt_moneda(flete_c))
                r1c4.metric("👤 Total Comisiones",   fmt_moneda(comision_c))

                r2c1, r2c2, r2c3, r2c4 = st.columns(4)
                r2c1.metric("🌙 Total Bonos",        fmt_moneda(bono_c))
                r2c2.metric("⛽ Comb. ($)",          fmt_moneda(comb_p_c))
                r2c3.metric("⛽ Comb. (gal)",        f"{round(comb_g_c,2)} gal")
                r2c4.metric("📈 Utilidad Neta",      fmt_moneda(utilidad_c))

                st.divider()
                try:
                    import plotly.express as px
                    gc1, gc2 = st.columns(2)
                    with gc1:
                        st.markdown("#### Viajes por mes")
                        df_cond_t["mes"] = pd.to_datetime(df_cond_t["fecha"]).dt.to_period("M").astype(str)
                        df_mes = df_cond_t.groupby("mes").size().reset_index(name="viajes")
                        fig_m = px.bar(df_mes, x="mes", y="viajes", text="viajes",
                                       color_discrete_sequence=["#e94560"])
                        fig_m.update_traces(textposition="outside")
                        fig_m.update_layout(margin=dict(t=10,b=10), height=260)
                        st.plotly_chart(fig_m, use_container_width=True)
                    with gc2:
                        st.markdown("#### Fletes por mes")
                        df_fmes = df_cond_t.groupby("mes")["flete"].sum().reset_index()
                        fig_fm = px.bar(df_fmes, x="mes", y="flete",
                                        color_discrete_sequence=["#2ecc71"], text="flete")
                        fig_fm.update_traces(texttemplate='$%{text:,.0f}', textposition="outside")
                        fig_fm.update_layout(margin=dict(t=10,b=10), height=260)
                        st.plotly_chart(fig_fm, use_container_width=True)
                except:
                    pass

                st.divider()
                st.markdown("#### 📋 Detalle de Viajes")
                cols_det = ["fecha","placa","cliente","origen","destino","estado",
                            "flete","comision","bono_transporte","combustible_pesos","combustible_galones","numero_factura"]
                cols_det_ex = [c for c in cols_det if c in df_cond_t.columns]
                df_det = df_cond_t[cols_det_ex].copy()
                df_det["utilidad"] = (
                    df_det.get("flete", pd.Series([0]*len(df_det))).fillna(0).astype(int)
                    - df_det.get("comision", pd.Series([0]*len(df_det))).fillna(0).astype(int)
                )
                st.dataframe(df_det, use_container_width=True, hide_index=True)

                excel_cond = generar_excel(df_cond_t, titulo=f"Trazabilidad - {conductor_sel_t}")
                st.download_button(
                    f"⬇️ Descargar Excel de {conductor_sel_t}",
                    data=excel_cond,
                    file_name=f"Trazabilidad_{conductor_sel_t}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info(f"No hay viajes registrados para **{conductor_sel_t}** en el período seleccionado.")

        with traza_tab2:
            placa_sel_t = st.selectbox("Seleccionar placa:", list(PLACA_CONDUCTOR.keys()), key="t_placa")

            with st.expander("📅 Filtro de fechas", expanded=False):
                tv1, tv2 = st.columns(2)
                with tv1: v_fi = st.date_input("Desde", datetime.now() - timedelta(days=90), key="v_fi")
                with tv2: v_ff = st.date_input("Hasta", datetime.now(), key="v_ff")

            df_veh_t = q_obtener_viajes(db, fecha_ini=v_fi, fecha_fin=v_ff, placa=placa_sel_t)

            if not df_veh_t.empty:
                total_v       = len(df_veh_t)
                flete_v       = int(df_veh_t["flete"].sum())
                comb_p_v      = int(df_veh_t["combustible_pesos"].sum())
                comb_g_v      = float(df_veh_t["combustible_galones"].sum())
                conductores_v = df_veh_t["conductor"].dropna().unique().tolist()

                conductor_veh_asoc = PLACA_CONDUCTOR.get(placa_sel_t, "—")
                st.markdown(f"""
                <div class="trazabilidad-header">
                    <h4>🚛 Placa: {placa_sel_t} — Conductor asociado: {conductor_veh_asoc}</h4>
                </div>
                """, unsafe_allow_html=True)

                vk1, vk2, vk3, vk4, vk5 = st.columns(5)
                vk1.metric("🚚 Viajes",              total_v)
                vk2.metric("💰 Fletes Generados",    fmt_moneda(flete_v))
                vk3.metric("⛽ Combustible ($)",      fmt_moneda(comb_p_v))
                vk4.metric("⛽ Combustible (gal)",    f"{round(comb_g_v,2)} gal")
                vk5.metric("👥 Conductores",          len(conductores_v))

                if len(conductores_v) > 1:
                    st.info(f"Este vehículo fue operado por: {', '.join(conductores_v)}")

                st.divider()
                st.markdown("#### 🗺️ Rutas operadas")
                df_rutas = df_veh_t.groupby(["origen","destino"]).agg(
                    viajes=("origen","count"),
                    flete_total=("flete","sum")
                ).reset_index().sort_values("viajes", ascending=False)
                df_rutas["ruta"] = df_rutas["origen"] + " → " + df_rutas["destino"]
                df_rutas["flete_fmt"] = df_rutas["flete_total"].apply(fmt_moneda)
                st.dataframe(df_rutas[["ruta","viajes","flete_fmt"]].rename(columns={
                    "ruta":"Ruta","viajes":"Viajes","flete_fmt":"Total Fletes"
                }), use_container_width=True, hide_index=True)

                st.divider()
                st.markdown("#### 📋 Historial completo del vehículo")
                cols_v = ["fecha","conductor","cliente","origen","destino","estado",
                          "flete","combustible_pesos","combustible_galones","numero_factura","observacion"]
                cols_v_ex = [c for c in cols_v if c in df_veh_t.columns]
                st.dataframe(df_veh_t[cols_v_ex], use_container_width=True, hide_index=True)

                excel_veh = generar_excel(df_veh_t, titulo=f"Trazabilidad - {placa_sel_t}")
                st.download_button(
                    f"⬇️ Descargar Excel de {placa_sel_t}",
                    data=excel_veh,
                    file_name=f"Trazabilidad_{placa_sel_t}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info(f"No hay viajes registrados para la placa **{placa_sel_t}** en el período seleccionado.")

    # ===================== TAB 5: CONTROL COMBUSTIBLE =====================
    with tab5:
        st.markdown("### ⛽ Control de Combustible por Vehículo")
        st.caption("Registra tanqueos y monitorea el saldo de galones de cada vehículo en tiempo real.")

        comb_tab1, comb_tab2 = st.tabs(["🛢️ Registrar Tanqueo", "📊 Saldo y Trazabilidad"])

        with comb_tab1:
            st.markdown("#### Nuevo Tanqueo")
            with st.form("form_tanqueo", clear_on_submit=True):
                t1, t2, t3 = st.columns(3)
                with t1:
                    t_fecha  = st.date_input("📅 Fecha tanqueo", datetime.now(), key="t_fecha")
                with t2:
                    t_placa  = st.selectbox("🚛 Placa", list(PLACA_CONDUCTOR.keys()), key="t_placa_reg")
                with t3:
                    cond_t   = PLACA_CONDUCTOR.get(t_placa, "")
                    t_cond   = st.selectbox("👤 Conductor", ["— Seleccionar —"] + TODOS_CONDUCTORES,
                                             index=(["— Seleccionar —"] + TODOS_CONDUCTORES).index(cond_t)
                                             if cond_t in TODOS_CONDUCTORES else 0, key="t_cond_reg")

                t4, t5, t6 = st.columns(3)
                with t4:
                    t_galones = st.number_input("🛢️ Galones tanqueados", min_value=0.1, step=0.5,
                                                 format="%.2f", key="t_gal")
                with t5:
                    t_costo_txt = st.text_input("💵 Costo total ($)", placeholder="Ej: 250.000", key="t_costo")
                with t6:
                    t_obs = st.text_input("📝 Observación (opcional)", key="t_obs")

                t_costo_int = parse_moneda(t_costo_txt)
                if t_costo_txt and t_costo_int > 0:
                    precio_gal = int(t_costo_int / t_galones) if t_galones > 0 else 0
                    st.caption(f"✅ {fmt_moneda(t_costo_int)} — Precio por galón: {fmt_moneda(precio_gal)}")

                sub_tanqueo = st.form_submit_button("💾 Registrar Tanqueo", type="primary", use_container_width=True)

            if sub_tanqueo:
                if t_galones <= 0:
                    st.error("⚠️ Los galones deben ser mayor a 0.")
                else:
                    datos_t = {
                        "fecha": t_fecha,
                        "placa": t_placa,
                        "conductor": "" if t_cond == "— Seleccionar —" else t_cond,
                        "galones": t_galones,
                        "costo_pesos": t_costo_int,
                        "observacion": t_obs,
                    }
                    if db.guardar_tanqueo(datos_t):
                        limpiar_cache()
                        st.success(f"✅ Tanqueo registrado — {t_placa} | {t_galones} gal | {fmt_moneda(t_costo_int)}")
                        st.rerun()

            st.divider()
            st.markdown("#### 📋 Historial de Tanqueos")
            with st.expander("Filtros tanqueos", expanded=False):
                ht1, ht2, ht3 = st.columns(3)
                with ht1: ht_placa = st.selectbox("Placa", ["Todas"] + list(PLACA_CONDUCTOR.keys()), key="ht_placa")
                with ht2: ht_fi    = st.date_input("Desde", datetime.now() - timedelta(days=90), key="ht_fi")
                with ht3: ht_ff    = st.date_input("Hasta", datetime.now(), key="ht_ff")

            df_tanqueos = q_obtener_tanqueos(
                db,
                placa=ht_placa if ht_placa != "Todas" else None,
                fecha_ini=ht_fi, fecha_fin=ht_ff
            )
            if not df_tanqueos.empty:
                st.dataframe(
                    df_tanqueos[["id","fecha","placa","conductor","galones","costo_pesos","observacion"]],
                    use_container_width=True, hide_index=True
                )
                st.markdown("**Eliminar tanqueo:**")
                tanqueo_ids = df_tanqueos["id"].tolist()
                tanqueo_labels = df_tanqueos.apply(
                    lambda r: f"ID {r['id']} | {r['fecha']} | {r['placa']} | {r['galones']} gal | {fmt_moneda(r['costo_pesos'])}",
                    axis=1
                ).tolist()
                sel_del_t = st.selectbox("Seleccionar tanqueo a eliminar:", tanqueo_labels, key="del_tanqueo_sel")
                if st.button("🗑️ Eliminar tanqueo seleccionado", key="btn_del_tanqueo"):
                    tid = int(sel_del_t.split(" | ")[0].replace("ID ", ""))
                    if db.eliminar_tanqueo(tid):
                        limpiar_cache()
                        st.success("Tanqueo eliminado."); st.rerun()
            else:
                st.info("No hay tanqueos registrados con esos filtros.")

        with comb_tab2:
            st.markdown("#### 📊 Saldo de Combustible por Vehículo")

            resumen_rows = []
            for placa_k in PLACA_CONDUCTOR.keys():
                saldo_info = q_saldo_combustible_placa(db, placa_k)
                resumen_rows.append({
                    "Placa": placa_k,
                    "Conductor": PLACA_CONDUCTOR[placa_k],
                    "Total Tanqueado (gal)": saldo_info["total_tanqueado"],
                    "Total Consumido (gal)": saldo_info["total_consumido"],
                    "Saldo Actual (gal)": saldo_info["saldo"],
                })
            df_resumen = pd.DataFrame(resumen_rows)

            st.markdown("##### Resumen general de flotilla")
            for _, row_r in df_resumen.iterrows():
                saldo = row_r["Saldo Actual (gal)"]
                cls   = "saldo-alto" if saldo > 20 else ("saldo-medio" if saldo > 5 else "saldo-bajo")
                icono = "🟢" if saldo > 20 else ("🟡" if saldo > 5 else "🔴")
                col_a, col_b, col_c, col_d, col_e = st.columns([2,2,2,2,3])
                col_a.write(f"**{row_r['Placa']}**")
                col_b.write(row_r["Conductor"])
                col_c.write(f"⬆️ {row_r['Total Tanqueado (gal)']} gal")
                col_d.write(f"⬇️ {row_r['Total Consumido (gal)']} gal")
                col_e.markdown(
                    f"{icono} Saldo: <span class='{cls}'>{saldo} gal</span>",
                    unsafe_allow_html=True
                )

            st.divider()
            st.markdown("##### Detalle por vehículo")
            placa_det = st.selectbox("Ver detalle de placa:", list(PLACA_CONDUCTOR.keys()), key="comb_det_placa")
            saldo_det = q_saldo_combustible_placa(db, placa_det)

            saldo_val = saldo_det["saldo"]
            cls_saldo = "saldo-alto" if saldo_val > 20 else ("saldo-medio" if saldo_val > 5 else "saldo-bajo")
            icono_s   = "🟢" if saldo_val > 20 else ("🟡" if saldo_val > 5 else "🔴")

            sd1, sd2, sd3 = st.columns(3)
            sd1.metric("🛢️ Total Tanqueado", f"{saldo_det['total_tanqueado']} gal")
            sd2.metric("🚛 Total Consumido", f"{saldo_det['total_consumido']} gal")
            sd3.markdown(
                f"<div class='combustible-card'>"
                f"<div style='font-size:0.72rem;color:#8892b0;text-transform:uppercase;letter-spacing:1px;'>Saldo Actual</div>"
                f"<div class='{cls_saldo}'>{icono_s} {saldo_val} gal</div>"
                f"</div>",
                unsafe_allow_html=True
            )

            if not saldo_det["historial"].empty:
                st.markdown("##### Historial de movimientos (tanqueos y consumos)")
                df_hist_show = saldo_det["historial"].copy()
                df_hist_show["tipo_icon"] = df_hist_show["tipo"].apply(
                    lambda x: "🛢️ TANQUEO" if x == "TANQUEO" else "🚛 CONSUMO"
                )
                df_hist_show["galones_fmt"] = df_hist_show.apply(
                    lambda r: f"+{r['galones']} gal" if r["tipo"] == "TANQUEO" else f"-{r['galones']} gal",
                    axis=1
                )
                df_hist_show["costo_fmt"] = df_hist_show["costo_pesos"].apply(fmt_moneda)
                cols_hist = ["fecha","tipo_icon","galones_fmt","costo_fmt","saldo_acumulado","observacion"]
                st.dataframe(
                    df_hist_show[cols_hist].rename(columns={
                        "fecha":"Fecha", "tipo_icon":"Tipo", "galones_fmt":"Galones",
                        "costo_fmt":"Costo ($)", "saldo_acumulado":"Saldo Acum. (gal)",
                        "observacion":"Referencia"
                    }),
                    use_container_width=True, hide_index=True
                )

                try:
                    import plotly.express as px
                    fig_saldo = px.line(
                        df_hist_show, x="fecha", y="saldo_acumulado",
                        title=f"Evolución del saldo de galones — {placa_det}",
                        markers=True,
                        color_discrete_sequence=["#2ecc71"]
                    )
                    fig_saldo.add_hline(y=10, line_dash="dot", line_color="#f39c12",
                                        annotation_text="Alerta 10 gal", annotation_position="bottom right")
                    fig_saldo.add_hline(y=5,  line_dash="dot", line_color="#e74c3c",
                                        annotation_text="Crítico 5 gal",  annotation_position="bottom right")
                    fig_saldo.update_layout(margin=dict(t=40,b=10), height=320,
                                            xaxis_title="Fecha", yaxis_title="Galones")
                    st.plotly_chart(fig_saldo, use_container_width=True)
                except:
                    pass
            else:
                st.info(f"No hay movimientos de combustible registrados para **{placa_det}**.")


if __name__ == "__main__":
    main()
